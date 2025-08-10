"""
Excel Processor Module
"""
import asyncio
import shutil
from datetime import timedelta
from typing import List, Dict
from openpyxl import load_workbook


class ExcelProcessor:
    """Process and fill Excel templates with ConEd data"""
    
    @staticmethod
    async def fill_template(template_path: str, output_path: str, data_points: List[Dict], 
                           username: str = None, account_id: str = None, region_code: str = None) -> int:
        """Fill Excel template with ConEd usage data"""
        
        # Copy template to output path
        print(f"\nCopying template to {output_path}...")
        shutil.copy2(template_path, output_path)
        print("✅ Template copied")
        
        # Open the workbook
        print(f"Opening Excel file: {output_path}")
        wb = load_workbook(output_path)
        ws = wb.active  # First sheet (Intervals)
        
        print(f"Sheet name: {ws.title}")
        
        # Fill in username and account ID
        if username:
            ws['B2'] = username
            print(f"Filled B2 with username: {username}")
        
        if account_id:
            ws['B4'] = account_id
            print(f"Filled B4 with account ID: {account_id}")
        
        # Set region code in RATE CALCULATIONS sheet cell B1
        if region_code:
            # Look for RATE CALCULATIONS sheet
            rate_calc_sheet = None
            for sheet_name in wb.sheetnames:
                if 'RATE CALCULATIONS' in sheet_name.upper():
                    rate_calc_sheet = wb[sheet_name]
                    break
            
            if rate_calc_sheet:
                rate_calc_sheet['B1'] = region_code
                print(f"Filled RATE CALCULATIONS B1 with region code: {region_code}")
            else:
                print(f"Warning: Could not find RATE CALCULATIONS sheet to set region code: {region_code}")
        
        # Sort data points by timestamp
        data_points.sort(key=lambda x: x['timestamp'])
        
        # Starting at row 8, fill in the usage data
        start_row = 8
        end_row = min(35033, start_row + len(data_points) - 1)
        
        print(f"Filling rows {start_row} to {end_row} with usage data...")
        
        filled_count = 0
        for i, data_point in enumerate(data_points):
            row = start_row + i
            if row > end_row:
                break
            
            # Remove timezone info for Excel compatibility
            timestamp_naive = data_point['timestamp'].replace(tzinfo=None)
            end_time_naive = data_point['end_time'].replace(tzinfo=None)
            
            # Subtract 1 minute from end time to match original format
            end_time_adjusted = end_time_naive - timedelta(minutes=1)
            
            # Column B: DATE
            ws.cell(row=row, column=2, value=timestamp_naive)
            
            # Column C: START TIME
            ws.cell(row=row, column=3, value=timestamp_naive.strftime('%H:%M:%S'))
            
            # Column D: END TIME  
            ws.cell(row=row, column=4, value=end_time_adjusted.strftime('%H:%M:%S'))
            
            # Column E: USAGE (kWh)
            ws.cell(row=row, column=5, value=data_point['usage_kwh'])
            
            filled_count += 1
            
            if filled_count % 1000 == 0:
                print(f"  Filled {filled_count} rows...")
            
            # Yield control every 5000 rows to allow WebSocket updates
            if filled_count % 5000 == 0:
                await asyncio.sleep(0)
        
        print(f"✅ Filled {filled_count} rows with data")
        
        # Save the file
        print(f"Saving to: {output_path}")
        wb.save(output_path)
        print(f"✅ Excel file saved successfully!")
        
        return filled_count