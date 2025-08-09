#!/usr/bin/env python3
"""
Script to fetch ConEd data and fill demo_sheet.xlsx with 15-minute interval data
"""
import asyncio
import aiohttp
import sys
import getpass
from pathlib import Path
from datetime import datetime, timedelta
import openpyxl
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent / "opower" / "src"))
from opower import Opower, AggregateType

async def fetch_coned_data(username, password, start_date, end_date):
    """
    Fetch 15-minute interval data from ConEd for the specified time period
    """
    async def mfa_callback():
        print("\n🔐 MFA Required!")
        mfa_code = input("Enter your 6-digit MFA code: ")
        return mfa_code
    
    data_points = []
    account_info = {}
    
    try:
        async with aiohttp.ClientSession() as session:
            print("Creating Opower API instance...")
            api = Opower(session, "coned", username, password, None)
            
            print("Starting login process...")
            await api.async_login(mfa_callback=mfa_callback)
            print("✅ Login successful!")
            
            print("\nFetching accounts...")
            accounts = await api.async_get_accounts()
            print(f"Found {len(accounts)} accounts")
            
            # Find electric account with quarter-hour resolution
            elec_account = None
            for account in accounts:
                if account.meter_type.value == 'ELEC' and account.read_resolution and 'QUARTER' in account.read_resolution.value:
                    elec_account = account
                    account_info['account_id'] = account.id
                    print(f"Using electric account: {account.id}")
                    break
            
            if not elec_account:
                raise Exception("No electric account with 15-minute interval capability found")
            
            print(f"\nFetching 15-minute interval data from {start_date.date()} to {end_date.date()}...")
            
            # Fetch data in chunks (API might have limits)
            current_start = start_date
            chunk_days = 30  # Fetch 30 days at a time
            
            while current_start < end_date:
                current_end = min(current_start + timedelta(days=chunk_days), end_date)
                
                print(f"Fetching chunk: {current_start.date()} to {current_end.date()}")
                
                usage_reads = await api.async_get_usage_reads(
                    elec_account,
                    AggregateType.QUARTER_HOUR,
                    start_date=current_start,
                    end_date=current_end
                )
                
                for reading in usage_reads:
                    data_points.append({
                        'timestamp': reading.start_time,
                        'end_time': reading.end_time,
                        'usage_kwh': reading.consumption
                    })
                
                print(f"  Fetched {len(usage_reads)} readings")
                current_start = current_end + timedelta(days=1)
            
            print(f"\n✅ Total data points fetched: {len(data_points)}")
            return data_points, account_info
            
    except Exception as e:
        print(f"❌ Error fetching data: {e}")
        import traceback
        traceback.print_exc()
        return None, None

def fill_excel_sheet(data_points, excel_file='demo_sheet.xlsx', output_file='demo_sheet_filled.xlsx', username=None, account_id=None):
    """
    Copy the Excel sheet and fill with the fetched ConEd data
    """
    import shutil
    
    # First, copy the original file
    print(f"\nCopying {excel_file} to {output_file}...")
    shutil.copy2(excel_file, output_file)
    print(f"✅ File copied")
    
    print(f"Opening Excel file: {output_file}")
    wb = load_workbook(output_file)
    ws = wb.active  # First sheet (Intervals)
    
    print(f"Sheet name: {ws.title}")
    
    # Fill in username and account ID
    if username:
        ws['B2'] = username
        print(f"Filled B2 with username: {username}")
    
    if account_id:
        ws['B4'] = account_id
        print(f"Filled B4 with account ID: {account_id}")
    
    # Sort data points by timestamp
    data_points.sort(key=lambda x: x['timestamp'])
    
    # Starting at row 8, fill in the usage data (column E)
    start_row = 8
    end_row = min(35033, start_row + len(data_points) - 1)
    
    print(f"Filling rows {start_row} to {end_row} with usage data...")
    
    filled_count = 0
    for i, data_point in enumerate(data_points):
        row = start_row + i
        if row > end_row:
            break
        
        # Remove timezone info for Excel compatibility
        timestamp_naive = data_point['timestamp'].replace(tzinfo=None)
        end_time_naive = data_point['end_time'].replace(tzinfo=None)
        
        # Subtract 1 minute from end time to match original format
        end_time_adjusted = end_time_naive - timedelta(minutes=1)
        
        # Column B: DATE (already filled, but we can verify/update)
        ws.cell(row=row, column=2, value=timestamp_naive)
        
        # Column C: START TIME
        ws.cell(row=row, column=3, value=timestamp_naive.strftime('%H:%M:%S'))
        
        # Column D: END TIME  
        ws.cell(row=row, column=4, value=end_time_adjusted.strftime('%H:%M:%S'))
        
        # Column E: USAGE (kWh)
        ws.cell(row=row, column=5, value=data_point['usage_kwh'])
        
        filled_count += 1
        
        if filled_count % 1000 == 0:
            print(f"  Filled {filled_count} rows...")
    
    print(f"✅ Filled {filled_count} rows with data")
    
    # Save the file
    print(f"Saving to: {output_file}")
    wb.save(output_file)
    print(f"✅ Excel file saved successfully!")
    
    # Reload workbook with data_only=True to get calculated values
    print("\nReloading workbook to get calculated values...")
    wb_calc = load_workbook(output_file, data_only=True)
    
    # Extract calculated costs from Rate Summary tab
    print("Extracting calculated costs from Rate Summary tab...")
    try:
        rate_summary_ws = wb_calc['RATE SUMMARY']
        costs = {
            'EL1': rate_summary_ws['F17'].value,
            'Time of Use': rate_summary_ws['I17'].value,
            'Smart Energy Plan': rate_summary_ws['L17'].value,
            'Select Pricing Plan': rate_summary_ws['O17'].value,
            'Standby': rate_summary_ws['R17'].value
        }
        
        print("\n📊 Calculated Costs (Row 17):")
        print("-" * 50)
        
        # Check if we got formula strings instead of values
        has_values = any(isinstance(cost, (int, float)) for cost in costs.values() if cost is not None)
        
        if not has_values:
            print("Note: Excel formulas need to be recalculated.")
            print("Open the file in Excel and save it to update formula values.")
            print("-" * 50)
            for rate_name, formula in costs.items():
                if formula is not None:
                    print(f"{rate_name:20} {formula}")
        else:
            for rate_name, cost in costs.items():
                if cost is not None:
                    print(f"{rate_name:20} ${cost:,.2f}" if isinstance(cost, (int, float)) else f"{rate_name:20} {cost}")
                else:
                    print(f"{rate_name:20} (no value)")
        print("-" * 50)
        
        return filled_count, costs
    except Exception as e:
        print(f"Warning: Could not extract rate costs: {e}")
        return filled_count, None

async def main():
    print("=" * 60)
    print("ConEd Data Fetcher - Fill Excel Sheet")
    print("=" * 60)
    
    # Get credentials
    username = input("Enter ConEd username: ")
    password = getpass.getpass("Enter ConEd password: ")
    
    # Define date range
    start_date = datetime(2024, 8, 1)
    end_date = datetime(2025, 7, 31, 23, 59, 59)
    
    print(f"\nDate range: {start_date.date()} to {end_date.date()}")
    print(f"Expected data points: ~{(end_date - start_date).days * 96} (96 per day)")
    
    # Step 1: Fetch data from ConEd
    print("\n" + "=" * 40)
    print("STEP 1: Fetching ConEd Data")
    print("=" * 40)
    
    data_points, account_info = await fetch_coned_data(username, password, start_date, end_date)
    
    if not data_points:
        print("❌ Failed to fetch data. Exiting.")
        return
    
    # Step 2: Fill Excel sheet
    print("\n" + "=" * 40)
    print("STEP 2: Filling Excel Sheet")
    print("=" * 40)
    
    account_id = account_info.get('account_id') if account_info else None
    result = fill_excel_sheet(data_points, username=username, account_id=account_id)
    
    # Handle both return formats
    if isinstance(result, tuple):
        filled_count, costs = result
    else:
        filled_count = result
        costs = None
    
    print("\n" + "=" * 60)
    print("✅ COMPLETE!")
    print(f"Fetched {len(data_points)} data points")
    print(f"Filled {filled_count} rows in Excel")
    print(f"Output saved to: demo_sheet_filled.xlsx")
    
    if costs:
        print("\n📊 RATE COMPARISON SUMMARY:")
        print("-" * 60)
        for rate_name, cost in costs.items():
            if cost is not None and isinstance(cost, (int, float)):
                print(f"{rate_name:20} ${cost:,.2f}")
    
    print("=" * 60)

if __name__ == "__main__":
    asyncio.run(main())